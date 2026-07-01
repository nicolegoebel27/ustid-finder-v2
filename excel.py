from openpyxl import load_workbook
from search import CompanySearcher


class ExcelProcessor:

    def __init__(self):
        self.searcher = CompanySearcher()

    def process_file(self, input_file, output_file):

        wb = load_workbook(input_file)
        ws = wb.active

        # Überschriften lesen
        headers = [cell.value for cell in ws[1]]

        # Neue Spalten anlegen
        ws.cell(row=1, column=len(headers) + 1).value = "USt-ID"
        ws.cell(row=1, column=len(headers) + 2).value = "Quelle"
        ws.cell(row=1, column=len(headers) + 3).value = "Status"

        # Spaltenpositionen suchen
        company_col = headers.index("Firma") + 1
        street_col = headers.index("Straße") + 1
        number_col = headers.index("Hnr.") + 1
        zip_col = headers.index("PLZ") + 1
        city_col = headers.index("Ort") + 1
        country_col = headers.index("Land") + 1

        # Alle Datenzeilen bearbeiten
        for row in range(2, ws.max_row + 1):

            firma = ws.cell(row, company_col).value or ""
            strasse = ws.cell(row, street_col).value or ""
            hnr = ws.cell(row, number_col).value or ""
            plz = ws.cell(row, zip_col).value or ""
            ort = ws.cell(row, city_col).value or ""
            land = ws.cell(row, country_col).value or ""

            result = self.searcher.search(
                firma=firma,
                street=strasse,
                number=hnr,
                zip_code=plz,
                city=ort,
                country=land
            )

            ws.cell(row, len(headers) + 1).value = result["vat"]
            ws.cell(row, len(headers) + 2).value = result["source"]
            ws.cell(row, len(headers) + 3).value = result["status"]

        wb.save(output_file)
