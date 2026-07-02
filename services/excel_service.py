from pathlib import Path

from openpyxl import load_workbook

from services.search_service import SearchService


class ExcelService:

    def __init__(self):
        self.search = SearchService()

    def process_file(self, input_path: Path, output_path: Path):

        wb = load_workbook(input_path)
        ws = wb.active

        # -------------------------
        # Überschriften einlesen
        # -------------------------

        headers = {}

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value

            if value:
                headers[str(value).strip()] = col

        required = [
            "Firma",
            "Straße",
            "Hnr.",
            "PLZ",
            "Ort",
            "Land"
        ]

        for field in required:
            if field not in headers:
                raise Exception(f"Spalte '{field}' wurde nicht gefunden.")

        # -------------------------
        # Neue Spalten anlegen
        # -------------------------

        website_col = ws.max_column + 1
        vat_col = ws.max_column + 2
        source_col = ws.max_column + 3
        status_col = ws.max_column + 4

        ws.cell(1, website_col).value = "Website"
        ws.cell(1, vat_col).value = "USt-ID"
        ws.cell(1, source_col).value = "Quelle"
        ws.cell(1, status_col).value = "Status"

        total = ws.max_row - 1

        print(f"Starte Verarbeitung von {total} Firmen...")

        # -------------------------
        # Firmen verarbeiten
        # -------------------------

        for row in range(2, ws.max_row + 1):

            try:

                firma = str(ws.cell(row, headers["Firma"]).value or "").strip()

                if not firma:
                    continue

                street = str(ws.cell(row, headers["Straße"]).value or "").strip()
                number = str(ws.cell(row, headers["Hnr."]).value or "").strip()
                zip_code = str(ws.cell(row, headers["PLZ"]).value or "").strip()
                city = str(ws.cell(row, headers["Ort"]).value or "").strip()
                country = str(ws.cell(row, headers["Land"]).value or "").strip()

                print(f"[{row-1}/{total}] {firma}")

                result = self.search.search_company(
                    company=firma,
                    street=street,
                    number=number,
                    zip_code=zip_code,
                    city=city,
                    country=country
                )

                ws.cell(row, website_col).value = result.get("website", "")
                ws.cell(row, vat_col).value = result.get("vat", "")
                ws.cell(row, source_col).value = result.get("source", "")
                ws.cell(row, status_col).value = result.get("status", "")

            except Exception as e:

                print(f"Fehler bei {firma}: {e}")

                ws.cell(row, website_col).value = ""
                ws.cell(row, vat_col).value = ""
                ws.cell(row, source_col).value = ""
                ws.cell(row, status_col).value = f"Fehler: {e}"

            # Alle 25 Firmen speichern
            if (row - 1) % 25 == 0:

                print(f"Zwischenspeichern... ({row-1}/{total})")

                wb.save(output_path)

        print("Endgültiges Speichern...")

        wb.save(output_path)

        print("Verarbeitung abgeschlossen.")
